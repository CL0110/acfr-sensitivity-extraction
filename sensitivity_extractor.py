"""
Sensitivity Analysis Extractor
Companion to Final.py — extracts discount rate sensitivity data from ACFR PDFs.

Pipeline: Ingest PDF → Identify sensitivity table pages → Send to Gemini → Parse 6 fields per plan → Write to Excel

Usage:
    python sensitivity_extractor.py --pdf-folder /path/to/acfr_pdfs --output sensitivity_results.xlsx
    python sensitivity_extractor.py --pdf-folder /path/to/acfr_pdfs --output sensitivity_results.xlsx --resume
"""

import os
import re
import csv
import json
import time
import argparse
import logging
from pathlib import Path
from difflib import SequenceMatcher

import PyPDF2
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ─────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────
API_KEY = ""  # Replace or use --api-key flag

# ─────────────────────────────────────────────
# State abbreviation → full name mapping
# ─────────────────────────────────────────────
STATE_ABBREV = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "DC": "District of Columbia",
}

# Common plan acronym → expanded name fragments (used as fallback matching hints)
PLAN_ACRONYM_HINTS = {
    "TRS": "teachers retirement",
    "ERS": "employees retirement",
    "PERS": "public employees retirement",
    "STRS": "state teachers retirement",
    "JRS": "judicial retirement",
    "JRF": "judicial retirement fund",
    "FRS": "florida retirement system",
    "SERS": "state employees retirement",
    "PSERS": "public school employees retirement",
    "PSPRS": "public safety personnel retirement",
    "ASRS": "arizona state retirement",
    "SCRS": "south carolina retirement",
    "KPERS": "public employees retirement",
    "IPERS": "public employees retirement",
    "NDPERS": "public employees retirement",
    "PERSI": "public employee retirement system of idaho",
    "VRS": "virginia retirement system",
    "WRS": "wisconsin retirement system",
    "SDRS": "south dakota retirement system",
    "MSRS": "state retirement",
    "MPSRS": "public school retirement",
    "MOSERS": "state employees retirement",
    "LASERS": "state employees retirement",
    "LEOFF": "law enforcement officers",
    "LAGERS": "local government",
    "NHRS": "new hampshire retirement",
    "MERS": "municipal employees retirement",
    "CJRS": "consolidated judicial retirement",
    "TSERS": "teachers and state employees retirement",
    "LGERS": "local governmental employees retirement",
    "CALPERS": "public employees retirement fund",
    "CALSTRS": "state teachers retirement fund",
    "OPERS": "public employees retirement",
    "STPF": "state patrol",
    # Arkansas-specific
    "ATRS": "teacher retirement system",
    "APERS": "public employees retirement system",
    "ARTRS": "teacher retirement system",
    # Alabama
    "RSA": "retirement system",  # covers ERS/TRS/JRF under RSA umbrella
    # Other common ones
    "PERA": "public employees retirement",
    "TPAF": "teachers pension and annuity",
    "PFRS": "police and firemen retirement",
    "SPRS": "state police retirement",
    "SHPRS": "state highway patrol retirement",
    "HPRS": "highway patrol",
    "ERS_HAZ": "employees retirement system hazardous",
    "ERS_NONHAZ": "employees retirement system nonhazardous",
    "CERS": "county employees retirement",
    "KTRS": "teachers retirement system",
    "MTRS": "teachers retirement system",
    "CTRS": "teachers retirement system",
    "OTRS": "teachers retirement system",
}


# ─────────────────────────────────────────────
# Filename parser
# ─────────────────────────────────────────────
def parse_filename(filename: str) -> dict:
    """Extract year, state abbreviation, and plan hint from a PDF filename.
    
    Handles common patterns like:
        2024_AL_RSA_ACFR.pdf
        AL_2024_ERS_ACFR.pdf
        2024-AL-TRS-ACFR.pdf
        Alabama_TRS_2024_ACFR.pdf
    """
    name = Path(filename).stem  # strip .pdf
    # Normalize separators
    parts = re.split(r'[_\-\s]+', name)
    
    result = {"year": None, "state_abbrev": None, "state_name": None, "plan_hint": None, "raw_parts": parts}
    
    for part in parts:
        upper = part.upper()
        # Year detection (2020-2029)
        if re.match(r'^20[2-9]\d$', part):
            result["year"] = part
        # State abbreviation (2-letter)
        elif upper in STATE_ABBREV and result["state_abbrev"] is None:
            result["state_abbrev"] = upper
            result["state_name"] = STATE_ABBREV[upper]
        # Skip known non-plan tokens
        elif upper in ("ACFR", "CAFR", "AFR", "FY", "PG", "PAGE", "FINAL", "DRAFT"):
            continue
        # Page number patterns
        elif re.match(r'^(pg|page)\d+', part, re.IGNORECASE):
            continue
        elif re.match(r'^\d+$', part) and len(part) <= 3:
            continue  # likely a page number
        # Everything else could be a plan identifier
        else:
            if result["plan_hint"] is None:
                result["plan_hint"] = upper
            else:
                result["plan_hint"] += " " + upper
    
    # Also try matching full state names in the filename
    if result["state_abbrev"] is None:
        name_lower = name.lower()
        for abbr, full_name in STATE_ABBREV.items():
            if full_name.lower().replace(" ", "") in name_lower.replace(" ", "").replace("_", "").replace("-", ""):
                result["state_abbrev"] = abbr
                result["state_name"] = full_name
                break
    
    return result


# ─────────────────────────────────────────────
# Plan list matcher
# ─────────────────────────────────────────────
class PlanMatcher:
    """Matches extracted plan names to the master plan list using state + fuzzy name matching."""
    
    def __init__(self, plan_list_path: str = None):
        """Load master plan list from a CSV or Excel file.
        Expected columns: YR, State, Plan Name
        """
        self.plans = []  # list of dicts: {"year": str, "state": str, "plan_name": str}
        
        if plan_list_path and os.path.exists(plan_list_path):
            ext = Path(plan_list_path).suffix.lower()
            if ext == ".csv":
                self._load_csv(plan_list_path)
            elif ext in (".xlsx", ".xls"):
                self._load_excel(plan_list_path)
            else:
                logging.warning(f"Unsupported plan list format: {ext}")
    
    def _load_csv(self, path):
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter="\t")  # tab-separated based on Claire's format
            for row in reader:
                self.plans.append({
                    "year": str(row.get("YR", "")).strip(),
                    "state": row.get("State", "").strip(),
                    "plan_name": row.get("Plan Name", "").strip(),
                })
        # Also try comma-separated if tab didn't work
        if not self.plans or not self.plans[0].get("state"):
            self.plans = []
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    self.plans.append({
                        "year": str(row.get("YR", "")).strip(),
                        "state": row.get("State", "").strip(),
                        "plan_name": row.get("Plan Name", "").strip(),
                    })
    
    def _load_excel(self, path):
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        yr_col = next((i for i, h in enumerate(headers) if h and "YR" in str(h).upper()), None)
        state_col = next((i for i, h in enumerate(headers) if h and "STATE" in str(h).upper()), None)
        name_col = next((i for i, h in enumerate(headers) if h and "PLAN" in str(h).upper()), None)
        
        if yr_col is None or state_col is None or name_col is None:
            logging.warning(f"Could not find YR/State/Plan Name columns in {path}")
            return
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.plans.append({
                "year": str(row[yr_col] or "").strip(),
                "state": str(row[state_col] or "").strip(),
                "plan_name": str(row[name_col] or "").strip(),
            })
        wb.close()
    
    def match(self, state_name: str, extracted_plan_name: str, year: str = None,
              filename_hint: str = None) -> dict | None:
        """Find the best match from the master list.
        
        Uses extracted_plan_name (from Gemini) as primary signal.
        Falls back to filename_hint (parsed from PDF filename) if Gemini name is weak.
        
        Returns {"year": ..., "state": ..., "plan_name": ..., "_match_score": ...} or None.
        """
        if not self.plans or not state_name:
            return None
        
        # Filter candidates by state (and optionally year)
        candidates = [p for p in self.plans if p["state"].lower() == state_name.lower()]
        if year:
            year_filtered = [p for p in candidates if p["year"] == year]
            if year_filtered:
                candidates = year_filtered
        
        if not candidates:
            return None
        
        # If only one plan for this state, return it directly
        if len(candidates) == 1 and not extracted_plan_name:
            return {**candidates[0], "_match_score": 1.0}
        
        # Try matching with Gemini-extracted name first, then filename hint as fallback
        search_terms = []
        if extracted_plan_name:
            search_terms.append(extracted_plan_name)
        if filename_hint and filename_hint.upper() != (extracted_plan_name or "").upper():
            search_terms.append(filename_hint)
        
        if not search_terms:
            if len(candidates) == 1:
                return {**candidates[0], "_match_score": 1.0}
            return None
        
        best_match = None
        best_score = 0.0
        
        for search_term in search_terms:
            term_lower = search_term.lower().strip()
            term_upper = search_term.upper().strip()
            
            # Get hint text for this term's acronym
            hint_text = PLAN_ACRONYM_HINTS.get(term_upper, "")
            
            for candidate in candidates:
                cand_lower = candidate["plan_name"].lower()
                
                # Build acronym from candidate name words
                cand_words = re.findall(r"[A-Za-z][a-z']*", candidate["plan_name"])
                cand_acronym = "".join(w[0].upper() for w in cand_words)
                
                # Also build acronym from significant words only (skip articles, prepositions)
                skip_words = {"of", "the", "and", "for", "in", "a", "an", "to"}
                cand_sig_acronym = "".join(
                    w[0].upper() for w in cand_words if w.lower() not in skip_words
                )
                
                # Direct acronym match
                if term_upper == cand_acronym or term_upper == cand_sig_acronym:
                    score = 0.95
                # Check if term contains the significant acronym or vice versa
                elif len(term_upper) > 2 and (cand_sig_acronym in term_upper or term_upper in cand_sig_acronym):
                    score = 0.88
                # Check if extracted name is a substring of candidate (or vice versa)
                elif term_lower in cand_lower or cand_lower in term_lower:
                    score = 0.90
                # Use the hint text for acronym-based matching
                elif hint_text and hint_text.lower() in cand_lower:
                    score = 0.85
                # Fuzzy match on the full name
                else:
                    score = SequenceMatcher(None, term_lower, cand_lower).ratio()
                
                # Slightly penalize matches from filename_hint vs direct Gemini extraction
                if search_term != extracted_plan_name and score < 0.95:
                    score *= 0.95
                
                if score > best_score:
                    best_score = score
                    best_match = candidate
        
        # Only return if confidence is reasonable
        if best_score >= 0.40:
            return {**best_match, "_match_score": round(best_score, 3)}
        
        return None

# Keywords that indicate a sensitivity analysis table is on the page
SENSITIVITY_KEYWORDS = [
    "1% decrease",
    "1% increase",
    "sensitivity of the net pension",
    "discount rate sensitivity",
    "net pension liability",
    "changes in the discount rate",
    "change in discount rate",
    "sensitivity of the net opeb",         # in case OPEB shows up nearby
    "current single discount rate",
    "current discount rate",
]

# The structured prompt sent to Gemini for each PDF
EXTRACTION_PROMPT = """You are a financial data extraction assistant. Your task is to extract the **discount rate sensitivity analysis** table from this public pension plan's Annual Comprehensive Financial Report (ACFR).

WHAT TO LOOK FOR:
- A table showing the Net Pension Liability (NPL) calculated at three discount rates:
  1. Current discount rate MINUS 1 percentage point
  2. Current discount rate
  3. Current discount rate PLUS 1 percentage point
- This table is required by GASB Statement No. 67/68 and appears in virtually every ACFR.
- The table header typically says something like "Sensitivity of the Net Pension Liability to Changes in the Discount Rate."
- Some ACFRs contain MULTIPLE pension plans (e.g., TRS, ERS, JRF). Extract a SEPARATE row for each plan.

IMPORTANT NOTES:
- Dollar amounts may be reported "in thousands" or "in millions" — note the unit but report the NUMBER AS PRINTED in the table (do not convert).
- Report rates as decimals (e.g., 0.0745 for 7.45%, or 7.45 if the table prints it as "7.45%").
- If the table shows rates as percentages like "7.45%", return the numeric value 7.45.
- If a plan shows a Net Pension ASSET (positive funded status) instead of a liability, still extract the values and note it.
- ONLY extract data from the sensitivity analysis table. Do NOT extract from other tables.
- If you find an OPEB (Other Post-Employment Benefits) sensitivity table, IGNORE it — only extract pension data.

RETURN FORMAT — respond with ONLY valid JSON, no markdown, no backticks, no explanation:
{
  "source_page": <page number where the table was found>,
  "dollar_unit": "<thousands|millions|ones>",
  "plans": [
    {
      "plan_name": "<plan name, e.g. 'TRS', 'ERS', 'PERS', 'STRS'>",
      "rateminus1": <discount rate minus 1%, as a number>,
      "current_discount_rate": <current discount rate, as a number>,
      "rateplus1": <discount rate plus 1%, as a number>,
      "nplminus1": <net pension liability at the lower rate, as a number>,
      "npl_current": <net pension liability at the current rate, as a number>,
      "nplplus1": <net pension liability at the higher rate, as a number>
    }
  ]
}

If there are multiple plans, include one object per plan in the "plans" array.
If you cannot find a sensitivity analysis table, return:
{"error": "No sensitivity analysis table found", "plans": []}
"""

# ─────────────────────────────────────────────
# Logging setup
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("sensitivity")


# ─────────────────────────────────────────────
# Step 1: Identify & extract relevant pages
# ─────────────────────────────────────────────
class SensitivityPageExtractor:
    """Extracts only pages containing the sensitivity analysis table."""

    def __init__(self, keywords=None):
        self.keywords = [k.lower() for k in (keywords or SENSITIVITY_KEYWORDS)]
        # Pre-compute space-stripped versions for broken-word matching
        # e.g. "discount rate" → "discountrate" to catch "dis count rate"
        self.keywords_stripped = [re.sub(r'\s+', '', kw) for kw in self.keywords]

    def _page_text(self, page) -> str:
        try:
            return (page.extract_text() or "").lower()
        except Exception:
            return ""

    def _keyword_matches(self, text: str) -> int:
        """Count how many keywords match the page text.
        
        Uses two passes to handle PDFs with broken word spacing:
        1. Normal match against the raw text
        2. Space-stripped match (removes ALL spaces from both text and keywords)
           to catch cases like "Dis count Rate" → "discountrate"
        """
        count = 0
        # Also prepare a space-stripped version of the full text
        text_stripped = re.sub(r'\s+', '', text)

        for kw, kw_stripped in zip(self.keywords, self.keywords_stripped):
            # Pass 1: normal substring match
            if kw in text:
                count += 1
            # Pass 2: space-stripped match (catches broken words)
            elif kw_stripped in text_stripped:
                count += 1

        return count

    def find_sensitivity_pages(self, pdf_path: str) -> list[int]:
        """Return 0-indexed page numbers that contain sensitivity keywords."""
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            hits = []
            for i, page in enumerate(reader.pages):
                text = self._page_text(page)
                # Require at least 2 keyword matches to reduce false positives
                match_count = self._keyword_matches(text)
                if match_count >= 2:
                    hits.append(i)
            return hits

    def extract_pages(self, input_path: str, output_path: str) -> tuple[list[int], int]:
        """Write a trimmed PDF with only the sensitivity pages.
        Returns (pages_extracted_indices, total_pages).
        """
        with open(input_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            total = len(reader.pages)
            pages = self.find_sensitivity_pages(input_path)

            if not pages:
                return [], total

            # Also grab the page before each hit (often has the table header/context)
            expanded = set()
            for p in pages:
                if p > 0:
                    expanded.add(p - 1)
                expanded.add(p)
                if p + 1 < total:
                    expanded.add(p + 1)

            writer = PyPDF2.PdfWriter()
            for p in sorted(expanded):
                writer.add_page(reader.pages[p])

            with open(output_path, "wb") as out:
                writer.write(out)

            return sorted(expanded), total


# ─────────────────────────────────────────────
# Step 2: Send to Gemini & parse response
# ─────────────────────────────────────────────
class GeminiExtractor:
    """Sends a trimmed PDF to Gemini and returns structured JSON."""

    def __init__(self, api_key: str, model_name: str = "gemini-2.5-flash"):
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel(model_name)

    def extract(self, pdf_path: str) -> dict:
        pdf_bytes = Path(pdf_path).read_bytes()
        pdf_part = {"mime_type": "application/pdf", "data": pdf_bytes}

        response = self.model.generate_content([EXTRACTION_PROMPT, pdf_part])
        text = response.text.strip()

        # Strip markdown fences
        if text.startswith("```json"):
            text = text[7:]
        if text.startswith("```"):
            text = text[3:]
        if text.endswith("```"):
            text = text[:-3]

        return json.loads(text.strip())


# ─────────────────────────────────────────────
# Step 3: Validate extracted data
# ─────────────────────────────────────────────
def validate_plan(plan: dict) -> list[str]:
    """Return a list of warnings (empty = all good)."""
    warnings = []

    rate_low = plan.get("rateminus1")
    rate_cur = plan.get("current_discount_rate")
    rate_high = plan.get("rateplus1")

    # Rate ordering check
    if all(v is not None for v in [rate_low, rate_cur, rate_high]):
        if not (rate_low < rate_cur < rate_high):
            warnings.append(f"Rate ordering unexpected: {rate_low} / {rate_cur} / {rate_high}")

        # Check the 1% spread (works for both decimal and percentage formats)
        diff_low = round(rate_cur - rate_low, 4)
        diff_high = round(rate_high - rate_cur, 4)
        if diff_low not in (0.01, 1.0):
            warnings.append(f"Low rate spread is {diff_low}, expected 0.01 or 1.0")
        if diff_high not in (0.01, 1.0):
            warnings.append(f"High rate spread is {diff_high}, expected 0.01 or 1.0")

    # NPL ordering check (higher rate → lower liability, generally)
    npl_low = plan.get("nplminus1")
    npl_cur = plan.get("npl_current")
    npl_high = plan.get("nplplus1")

    if all(v is not None for v in [npl_low, npl_cur, npl_high]):
        if not (npl_low >= npl_cur >= npl_high):
            warnings.append(f"NPL ordering unexpected: {npl_low} / {npl_cur} / {npl_high}")

    return warnings


# ─────────────────────────────────────────────
# Step 4: Write results to Excel
# ─────────────────────────────────────────────
def write_to_excel(all_results: dict, output_path: str, plan_matcher: PlanMatcher = None):
    """Write all extracted sensitivity data to a clean Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sensitivity Analysis"

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "PDF Filename",
        "YR",
        "State",
        "Plan Name (Master List)",
        "Plan Name (Extracted)",
        "Match Score",
        "Rate Minus 1%",
        "Current Discount Rate",
        "Rate Plus 1%",
        "NPL at Lower Rate",
        "NPL at Current Rate",
        "NPL at Higher Rate",
        "Dollar Unit",
        "Source Page",
        "Validation Warnings",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    row = 2
    for pdf_name, result in sorted(all_results.items()):
        # Parse filename for state/year
        parsed = parse_filename(pdf_name)
        state_name = parsed.get("state_name", "")
        year = parsed.get("year", "")

        if "error" in result and not result.get("plans"):
            # Write error row
            ws.cell(row=row, column=1, value=pdf_name).border = thin_border
            ws.cell(row=row, column=2, value=year).border = thin_border
            ws.cell(row=row, column=3, value=state_name).border = thin_border
            ws.cell(row=row, column=15, value=result.get("error", "Unknown error")).border = thin_border
            ws.cell(row=row, column=15).font = Font(color="FF0000")
            row += 1
            continue

        for plan in result.get("plans", []):
            warnings = validate_plan(plan)
            extracted_name = plan.get("plan_name", "")

            # Attempt match to master list
            matched = None
            if plan_matcher and plan_matcher.plans:
                matched = plan_matcher.match(
                    state_name, extracted_name, year,
                    filename_hint=parsed.get("plan_hint", "")
                )

            ws.cell(row=row, column=1, value=pdf_name).border = thin_border
            ws.cell(row=row, column=2, value=year).border = thin_border
            ws.cell(row=row, column=3, value=state_name).border = thin_border
            ws.cell(row=row, column=4, value=matched["plan_name"] if matched else "").border = thin_border
            ws.cell(row=row, column=5, value=extracted_name).border = thin_border

            match_score = matched.get("_match_score", "") if matched else ""
            score_cell = ws.cell(row=row, column=6, value=match_score)
            score_cell.border = thin_border
            if match_score and match_score < 0.80:
                score_cell.font = Font(color="FF6600")  # orange = low confidence match

            ws.cell(row=row, column=7, value=plan.get("rateminus1")).border = thin_border
            ws.cell(row=row, column=8, value=plan.get("current_discount_rate")).border = thin_border
            ws.cell(row=row, column=9, value=plan.get("rateplus1")).border = thin_border
            ws.cell(row=row, column=10, value=plan.get("nplminus1")).border = thin_border
            ws.cell(row=row, column=11, value=plan.get("npl_current")).border = thin_border
            ws.cell(row=row, column=12, value=plan.get("nplplus1")).border = thin_border
            ws.cell(row=row, column=13, value=result.get("dollar_unit", "")).border = thin_border
            ws.cell(row=row, column=14, value=result.get("source_page", "")).border = thin_border

            warn_text = "; ".join(warnings) if warnings else ""
            warn_cell = ws.cell(row=row, column=15, value=warn_text)
            warn_cell.border = thin_border
            if warnings:
                warn_cell.font = Font(color="FF6600")

            # Number formatting
            for col in (7, 8, 9):
                ws.cell(row=row, column=col).number_format = '0.00%' if plan.get("current_discount_rate", 0) < 1 else '0.00"%"'
            for col in (10, 11, 12):
                ws.cell(row=row, column=col).number_format = '#,##0.00'

            row += 1

    # Auto-fit column widths (approximate)
    col_widths = [40, 8, 18, 45, 25, 12, 16, 20, 16, 22, 22, 22, 12, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    log.info(f"Excel saved → {output_path}")


# ─────────────────────────────────────────────
# Main pipeline
# ─────────────────────────────────────────────
def run_pipeline(pdf_folder: str, output_xlsx: str, api_key: str, resume: bool = False, plan_list_path: str = None):
    """Full pipeline: clean → extract → validate → match to plan list → write Excel."""

    # Paths
    pdf_folder = os.path.abspath(pdf_folder)
    folder_name = os.path.basename(pdf_folder)
    trimmed_folder = os.path.join(os.path.dirname(pdf_folder), f"Sensitivity_Trimmed_{folder_name}")
    json_cache = os.path.join(os.path.dirname(pdf_folder), "sensitivity_cache.json")
    Path(trimmed_folder).mkdir(parents=True, exist_ok=True)

    # Load plan list for matching
    plan_matcher = PlanMatcher(plan_list_path)
    if plan_matcher.plans:
        log.info(f"Loaded {len(plan_matcher.plans)} plans from master list")
    else:
        log.warning("No plan list loaded — output will not include matched plan names")

    # Load cache if resuming
    if resume and os.path.exists(json_cache):
        with open(json_cache, "r", encoding="utf-8") as f:
            results = json.load(f)
        log.info(f"Resumed from cache — {len(results)} PDFs already processed")
    else:
        results = {}

    # Discover PDFs
    pdf_files = sorted([f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")])
    log.info(f"Found {len(pdf_files)} PDFs in {pdf_folder}")

    extractor_pages = SensitivityPageExtractor()
    gemini = GeminiExtractor(api_key=api_key)

    consecutive_errors = 0
    MAX_CONSECUTIVE_ERRORS = 3

    for i, pdf_file in enumerate(pdf_files, 1):
        # Skip if already cached successfully
        if pdf_file in results and "error" not in results[pdf_file]:
            log.info(f"[{i}/{len(pdf_files)}] Skipping (cached): {pdf_file}")
            continue

        log.info(f"[{i}/{len(pdf_files)}] Processing: {pdf_file}")

        input_path = os.path.join(pdf_folder, pdf_file)
        trimmed_path = os.path.join(trimmed_folder, pdf_file)

        # Step 1: Identify and extract sensitivity pages
        try:
            pages_found, total_pages = extractor_pages.extract_pages(input_path, trimmed_path)
            if not pages_found:
                log.warning(f"  No sensitivity table pages found — skipping")
                results[pdf_file] = {"error": "No sensitivity keywords found on any page", "plans": []}
                _save_cache(results, json_cache)
                continue
            log.info(f"  Found sensitivity content on pages {[p+1 for p in pages_found]} of {total_pages}")
        except Exception as e:
            log.error(f"  PDF reading error: {e}")
            results[pdf_file] = {"error": f"PDF read error: {e}", "plans": []}
            _save_cache(results, json_cache)
            continue

        # Step 2: Send to Gemini
        try:
            extracted = gemini.extract(trimmed_path)
            results[pdf_file] = extracted
            consecutive_errors = 0

            plan_count = len(extracted.get("plans", []))
            log.info(f"  ✓ Extracted {plan_count} plan(s)")

            # Inline validation
            for plan in extracted.get("plans", []):
                warnings = validate_plan(plan)
                if warnings:
                    for w in warnings:
                        log.warning(f"    ⚠ {plan.get('plan_name', '?')}: {w}")

        except Exception as e:
            consecutive_errors += 1
            log.error(f"  ✗ API error ({consecutive_errors}/{MAX_CONSECUTIVE_ERRORS}): {e}")
            results[pdf_file] = {"error": f"API error: {e}", "plans": []}

            if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                log.critical(f"Stopping — {MAX_CONSECUTIVE_ERRORS} consecutive API errors")
                _save_cache(results, json_cache)
                break

        # Save cache after each PDF
        _save_cache(results, json_cache)

        # Rate limiting — Gemini free tier is 15 RPM for flash
        time.sleep(4)

    # Step 4: Write to Excel
    write_to_excel(results, output_xlsx, plan_matcher=plan_matcher)

    # Summary
    total = len(results)
    success = sum(1 for r in results.values() if r.get("plans"))
    errors = total - success
    total_plans = sum(len(r.get("plans", [])) for r in results.values())
    log.info(f"\nDone! {success}/{total} PDFs extracted successfully ({total_plans} plan rows), {errors} errors")
    log.info(f"Results → {output_xlsx}")
    log.info(f"Cache   → {json_cache}")


def _save_cache(results: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)


# ─────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract sensitivity analysis data from ACFR PDFs")
    parser.add_argument("--pdf-folder", required=True, help="Folder containing ACFR PDFs")
    parser.add_argument("--output", default="sensitivity_results.xlsx", help="Output Excel filename")
    parser.add_argument("--api-key", default=API_KEY, help="Gemini API key")
    parser.add_argument("--resume", action="store_true", help="Resume from cached results")
    parser.add_argument("--plan-list", default=None,
                        help="Path to master plan list (CSV or Excel with YR/State/Plan Name columns)")
    args = parser.parse_args()

    run_pipeline(
        pdf_folder=args.pdf_folder,
        output_xlsx=args.output,
        api_key=args.api_key,
        resume=args.resume,
        plan_list_path=args.plan_list,
    )
